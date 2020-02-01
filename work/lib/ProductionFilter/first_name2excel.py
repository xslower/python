import pandas as pd
import os

print("Running from", os.path.dirname(os.path.realpath(__file__)))

data_name = "tb_cate_modified.csv"
path_data = os.path.join(os.path.dirname(os.path.realpath(__file__)), "data/" + data_name)

data = pd.read_csv(path_data, encoding="utf8")
# 把 nan 数据替换为 None
data = data.where(data.notnull(), None)


first_name = []
for i in data["first_name"]:
    if i:
        if i in first_name:
            continue
        else:
            first_name.append(i)


first_id = []
for i in data["first_cid"]:
    if i:
        if i in first_id:
            continue
        else:
            first_id.append(i)


from openpyxl import Workbook
wb = Workbook()
ws = wb.worksheets[0]
ws.title = "first sheet"
ws.cell(row=1, column=1, value="first_name")
ws.cell(row=1, column=2, value="first_cid")
ws.cell(row=1, column=3, value="second_name")
for i in range(len(first_name)):
    str = ""
    ws.cell(row=i + 2, column=1, value=first_name[i])
    ws.cell(row=i + 2, column=2, value=first_id[i])

    a = data.loc[data["first_name"] == first_name[i], "second_name"]
    second_name = []
    for j in a:
        if j:
            if j in second_name:
                continue
            else:
                second_name.append(j)
    for sec_name in second_name:
        str += sec_name.strip()
        str += "、"
    str = str[:-1]
    ws.cell(row=i + 2, column=3, value=str)

wb.save(filename='first_name.xlsx')
